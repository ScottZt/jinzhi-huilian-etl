import pandas as pd
import openpyxl
from typing import Optional, Dict, Any, List
from pathlib import Path


class ExcelAdapter:

    def __init__(self, config: Dict[str, Any]):
        self.file_path = config.get("file_path")
        self.sheet_name = config.get("sheet_name", 0)
        self.has_header = config.get("has_header", True)

    def read_excel(self, file_path: Optional[str] = None, nrows: Optional[int] = None) -> pd.DataFrame:
        path = file_path or self.file_path
        if not path:
            raise ValueError("file_path is required")

        header = 0 if self.has_header else None
        df = pd.read_excel(path, sheet_name=self.sheet_name, header=header, nrows=nrows)
        return df

    def get_file_info(self, file_path: Optional[str] = None) -> Dict[str, Any]:
        path = Path(file_path or self.file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        wb = openpyxl.load_workbook(str(path), read_only=True)
        sheet_names = wb.sheetnames
        first_sheet = wb[sheet_names[0]]
        row_count = first_sheet.max_row or 0
        col_count = first_sheet.max_column or 0
        wb.close()

        df_sample = pd.read_excel(str(path), nrows=5)

        return {
            "file_path": str(path),
            "file_size": path.stat().st_size,
            "sheets": sheet_names,
            "total_rows": row_count - 1,
            "total_columns": col_count,
            "columns": df_sample.columns.tolist(),
            "sample_data": df_sample.head(5).to_dict(orient='records'),
        }

    def check_connectivity(self) -> tuple[bool, str]:
        try:
            path = Path(self.file_path)
            if not path.exists():
                return False, f"File not found: {self.file_path}"
            wb = openpyxl.load_workbook(str(path), read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return True, f"Excel file with {len(sheets)} sheets: {', '.join(sheets)}"
        except Exception as e:
            return False, str(e)
